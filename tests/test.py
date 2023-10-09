#-------------------------------------------------------------  gen otp with time in hash key encode and decode  --------------------------------------------------------------------------
# from datetime import datetime, timedelta
# from pytz import timezone
# import base64,hashlib,secrets

# def hash_key():
#     RANDOM_STRING_CHARS = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
#     string  = ''.join(secrets.choice(RANDOM_STRING_CHARS) for i in range(12))
#     return hashlib.sha1(string.encode('utf-8')).hexdigest()[:6]

# def generate_otp(otp):
#     created_time = datetime.now(timezone('UTC'))
#     seconds = created_time.timestamp()
#     key = str(seconds) + '~' + otp
#     activation_key = base64.urlsafe_b64encode(
#         key.encode('ascii')).decode('utf-8')
#     return activation_key

# def decode(encoded):
#     if len(encoded) % 4:
#         encoded += '=' * (4 - len(encoded) % 4)
#     # decoded = base64.b64decode(encoded.encode('ascii')).decode('utf-8')
#     decoded = base64.urlsafe_b64decode(
#         encoded.encode('ascii')).decode('utf-8')
#     if ',' in decoded:
#         return decoded.split(',')
#     return decoded.split('~')


# def verify_email_link(key):
#     key = decode(key)
#     ts = float(key[0])
#     expire_datetime = datetime.fromtimestamp(ts)
#     if ((expire_datetime + timedelta(
#             minutes=8)) < datetime.now()):
#         msg = 'Email Varification key is expired'
#         # logger.error(str(key[3]) + " - " + msg)
#         return False #[False, msg]
#     return True #[True, 'success']

# otp = hash_key()
# print(otp)
# key=generate_otp(otp)
# print(key)
# decod = decode(key)
# print(decod)
# save_otp=decod[1]
# key_='MTY5MDMwOTk4NS4wNDIwNjh-YzdkZWQ1'
# print(verify_email_link(key_))

#---------------------------------------------------------------------------   end    ----------------------------------------------------------------------------


#---------------------------------------------------------------------import party and vehicle test ----------------------------------------------------------------------------

# Party Type (Client/Vendor)                     object
# Category (General/Fuel/Vehicle Provider)       object
# Company Name                                   object
# Email                                          object
# Display Name                                   object
# Contact No country code                       float64
# Contact Number                                 object
# VAT Registration (Registered/Unregistered)    float64
# TRN                                            object 
import json
from datetime import datetime
import os
from openpyxl import load_workbook
# import pandas as pd
# from models import Party
mobile_codes = ['93', '35818', '355', '213', '1684', '376', '244', '1264', '1268', '54', '374', '297', '61', '43', '994', '1242', '973', '880', '1246', '375', '32', '501', '229', '1441', '975',
                '591', '599', '387', '267', '0055', '55', '246', '673', '359', '226', '257', '855', '237', '1', '238', '1345', '236', '235', '56', '86', '61', '61', '57', '269', '242', '243',
                  '682', '506', '225', '385', '53', '599', '357', '420', '45', '253', '1767', '1809',' 1829', '670', '593', '20', '503', '240', '291', '372', '251', '500', '298', '679',
                    '358', '33', '594', '689','241', '220', '995', '49', '233', '350', '30', '299', '1473', '590', '1671', '502', '441481', '224', '245', '592', '509', '', '504',
                      '852', '36', '354', '91', '62', '98', '964', '353', '972', '39', '1876', '81', '441534', '962', '7', '254', '686', '850', '82', '383', '965', '996', '856',
                        '371', '961', '266', '231', '218', '423', '370', '352', '853', '389', '261', '265', '60', '960', '223', '356', '441624', '692', '596', '222', '230', '262',
                          '52', '691', '373', '377', '976', '382', '1664', '212', '258', '95', '264', '674', '977', '31', '687', '64', '505', '227', '234', '683', '672', '1670',
                            '47', '968', '92', '680', '970', '507', '675', '595', '51', '63', '870', '48', '351', '1787',' 1939', '974', '262', '40', '7', '250', '290', '1869'
                            , '1758', '508', '1784', '590', '590', '685', '378', '239', '966', '221', '381', '248', '232', '65', '1721', '421', '386', '677', '252', '27', '',
                              '211', '34', '94', '249', '597', '47', '268', '46', '41', '963', '886', '992', '255', '66', '228', '690', '676', '1868', '216', '90', '993', '1649',
                                '688', '256', '380', '971', '44', '1', '1', '598', '998', '678', '379', '58', '84', '1284', '1340', '681', '212', '967', '260', '263']

def excel_data_import1():
    file_path = '/home/ayushgurjar/Downloads/DSA-01B/PartyImport1 .xlsx'
    if not os.path.exists(file_path):
        print("File not found at the specified path")
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        # df = pd.read_excel(file_path)
        # df2 =  df.fillna("")
        # print(df.dtypes) #.to_json()
    except Exception as e:
        print(f'Error reading the Excel file {str(e)}')

    file = []
    total = 0
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == None:
                break
            # this code for party type--------------------------
            party_type = row[0]
            if party_type.lower() == "client":
                party_type = 0
            elif party_type.lower() == "vendor":
                party_type = 1
            else:
                party_type = -1
            # print(row[0],party_type)
            #this code for vender party type--------------------General/Fuel/Vehicle Provider
            vender_type = row[1]
            if vender_type.lower() == "general":
                vender_type = 0
            elif vender_type.lower() == "fuel":
                vender_type = 1
            else:
                vender_type = -1
            # print(row[1],vender_type)


            ##-------this code for mobile code and number-------
            m_code = "971"
            mob = row[6]
            if mob==None:
                mob = ''
                m_code = ''
            if type(mob)==float:
                mob = str(int(mob))
            if type(mob)==int:
                mob=str(mob)
            if type(mob)==str:
                mob = mob.replace('=','').replace('+','').replace('-','').replace(' ','')
                mob = mob.split('/')[0].split('&')[0] 
            if mob[:2]=='00':
                mob=mob[2:]
            code = mob[:3]
            m = mob
            if code in mobile_codes:
                m_code,mob=code,mob[3:]
            if m_code != '':    
                m_code = "+"+m_code
            # print("col-",total+1,"  ",m_code,",",mob,",",)

            #---------this code for gmail --------------------
            gmail=''
            if row[3]!=None:
                gmail = row[3]
                gmail=gmail.lower()
                if gmail=="@gmail.com":
                    gmail = ''

            # print(row[3],"-",gmail)
            #------------this code for TRN-----------
            if row[8]==None:
                trn = ''
            elif type(row[8])==str:
                trn = row[8]
                if trn.isalnum():
                    trn = trn
                else:
                    trn = trn.replace('/','').split()
                    for i in trn:
                        if i.isalnum():
                            trn = i
                        else:
                            trn=''
                if trn.isalpha():
                    trn=''
            elif type(row[8])==int:
                trn = str(row[8])
            elif type(row[8])==float:
                trn = str(int(row[8]))
            tx=trn
            if len(trn)<=10 and trn!="":
                trn = ""
            if len(trn)>15:
                trn =""
                print(trn)
            print(total,trn,"--",row[8],len(tx))

            
            # print(type(trn),trn)


            data = {
                'party_type': party_type,
                'vendor_party_type' : vender_type,
                'company_name': row[2],                              
                'email_address' : gmail,                                 
                'display_name' : row[4],                             
                'mobile_country_code' : m_code,                  
                'mobile' : mob,                           
                'VAT Registration' : row[7] if row[7] else '',  #none to empty string
                'TRN' : row[8] if row[8] else ''
            }
            file.append(data)
            total+=1
    except Exception as e:
        print(f'Error importing data into the database: {str(e)}')

    # print(file)
    # filename = "party.json"
    # with open(filename, "w") as f:
    #     json.dump(file, f, indent=4)
    print('Data imported successfully',str(total))

# excel_data_import1()

#--------------------------------------------------------------------------------

# Vehicle Type (Own/Market)                     object
# Company Name                                  object
# Vehicle Make (for "Own" vehicle type)         object
# Vehicle Model                                 object
# Vehicle Number                                object
# Vehicle Chassis Number                        object
# Vehicle Make Year (for "Own" vehicle type)     int64
# Vehicle Registration Date                     object
# Vehicle Registration Expiry Date              object

   ## create owner_type----- don't use-----------------------------
   ## label_o = ""
   #obj_o_type,_o = DropdownOptionValues.objects.get_or_create(label=label_o,tenant=request.tenant,
   #                                            defaults={"tenant":request.tenant,
   #                                            "key":"vehicle-owner-type","label":label_o,
   #                                            "created_by":request.user})
   ## print(obj_o_type,_o)
                    

# import os
# import pandas as pd
# from models import Vehicle
# from openpyxl import load_workbook
def excel_data_import2():
    file_path = '/home/ayushgurjar/Downloads/DSA-01B/Vehicle_Import.xlsx'

    if not os.path.exists(file_path):
        print("File not found at the specified path")

    try:
        # df = pd.read_excel(file_path)
        # df2 =  df.fillna("")
        # print(df.dtypes) #.to_json()
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        # print(sheet)
    except Exception as e:
        print(f'Error reading the Excel file {str(e)}')

    file = []
    total = 0
    try:
        for row in sheet.iter_rows(min_row=2,values_only=True):
            if row[0] == None:
                break
            # ------this code for vehicle num,type----------
            mix = row[4].split()
            num = mix[0]
            v_type = " ".join(mix[1:])

            #----------------this code for Vehicle Make Year----------------------
            make_year = str(int(row[6]))
            # print(make_year)

            #----------------this code for Registration Expiry Date----------------------
            if type(row[7])==str:
                exp_date = row[7].replace("-","/")
                if len(exp_date) == 10:
                    exp_date = datetime.strptime(exp_date, '%d/%m/%Y').date()#.strftime('%d/%m/%Y')
                else:
                    exp_date = datetime.strptime(exp_date, '%d/%m/%y').date()#.strftime('%d/%m/%Y')
            else:
                exp_date= (row[7]).date()#.strftime('%d/%m/%Y')

            # print("col",total+1,"-", exp_date,"-",type(exp_date))

            #----------------this code for Registration Date----------------------           
            r_date = (row[8])
            if type(r_date)==str:
                r_date = r_date.replace("-","/")
                if len(r_date) == 10:
                    r_date = datetime.strptime(r_date, '%d/%m/%Y').date()
                else:
                    r_date = datetime.strptime(r_date, '%d/%m/%y').date()
            else:
                r_date= r_date.date()
            
            # print("col",total+1,"-", exp_date,"-",type(exp_date))
            # print("col",total+1,"-",r_date , exp_date)     
            if not r_date<exp_date:
                r_date,exp_date = exp_date,r_date

            print(r_date<exp_date)
            
            data = {
                    'owner_type' : row[0],               
                    'Company Name' : row[1],                              
                    'Vehicle Make' : row[2],    
                    'Vehicle Model' : row[3] if row[3] else '',                             
                    'reg_number' : num,   
                    'vehicle_type': v_type,                        
                    'chassis_number' : row[5],                  
                    'Vehicle Make_Year' : make_year,
                    'issue_date' : r_date,          
                    'expiry_date' : exp_date,          
            }
            # file.append(data)
            total+=1
    except Exception as e:
        print(f'Error importing data into the database: {str(e)}')

    # print(file)
    # filename = "vehicle.json"
    # with open(filename, "w") as f:
    #     json.dump(file, f, indent=4)
    i = 0
    for veh in file:
            owner = veh["Company Name"]
            chassis_number = veh["chassis_number"]
            reg_number = veh["reg_number"]
            make_name = veh["Vehicle Make"]
            model_name= veh["Vehicle Model"]
            label_v = veh["vehicle_type"]
            label_o = veh["owner_type"] # not ib use
            issue_date = veh["issue_date"]
            expiry_date = veh["expiry_date"]
            i+=1
            print(i,owner,chassis_number,reg_number,make_name,model_name,label_o,label_v,issue_date,expiry_date)
            # print(veh)
    print('Data imported successfully',str(total))

# excel_data_import2()

#----------------------------------------------------------------------------------------------------------------

# file ="/home/ayushgurjar/Documents/TransportSimple/transportsimple-server/transportsimple/base/data_dump/countries.json"
# with open(file,'r') as f:
#     codes = json.load(f)
#     list = []
#     for data in codes:
#         code = data['phone_code'].replace('+','').replace('-','')
#         list.append(code)
#     dict = list

##############################################################################final import vehicle and party##########################################################################
                                          ############vehicle###################
# from django.db import transaction
# from transportsimple.base.exception_handler import FormError, InvalidDataError, handle_exceptions
# from transportsimple.base.service_status import created, deleted, updated, get
# from transportsimple.company.repo.tenant.models import TenantUser
# from transportsimple.parties.repo.detail.models import Party,  PartyContact
# from transportsimple.company.repo.document.models import Document
# from django.contrib.contenttypes.models import ContentType
# from transportsimple.company.constants import DocumentBelongsTo
# from transportsimple.tax.service import party_tax_fetcher, party_tax_creator
# from transportsimple.company.repo.address.models import Address
# import os
# import json
# from django.db import transaction
# from openpyxl import load_workbook
# from transportsimple.base.messages import *


# mobile_codes = ['93', '35818', '355', '213', '1684', '376', '244', '1264', '1268', '54', '374', '297', '61', '43', '994', '1242', '973', '880', '1246', '375', '32', '501', '229', '1441', '975',
#                 '591', '599', '387', '267', '0055', '55', '246', '673', '359', '226', '257', '855', '237', '1', '238', '1345', '236', '235', '56', '86', '61', '61', '57', '269', '242', '243',
#                   '682', '506', '225', '385', '53', '599', '357', '420', '45', '253', '1767', '1809',' 1829', '670', '593', '20', '503', '240', '291', '372', '251', '500', '298', '679',
#                     '358', '33', '594', '689','241', '220', '995', '49', '233', '350', '30', '299', '1473', '590', '1671', '502', '441481', '224', '245', '592', '509', '', '504',
#                       '852', '36', '354', '91', '62', '98', '964', '353', '972', '39', '1876', '81', '441534', '962', '7', '254', '686', '850', '82', '383', '965', '996', '856',
#                         '371', '961', '266', '231', '218', '423', '370', '352', '853', '389', '261', '265', '60', '960', '223', '356', '441624', '692', '596', '222', '230', '262',
#                           '52', '691', '373', '377', '976', '382', '1664', '212', '258', '95', '264', '674', '977', '31', '687', '64', '505', '227', '234', '683', '672', '1670',
#                             '47', '968', '92', '680', '970', '507', '675', '595', '51', '63', '870', '48', '351', '1787',' 1939', '974', '262', '40', '7', '250', '290', '1869'
#                             , '1758', '508', '1784', '590', '590', '685', '378', '239', '966', '221', '381', '248', '232', '65', '1721', '421', '386', '677', '252', '27', '',
#                               '211', '34', '94', '249', '597', '47', '268', '46', '41', '963', '886', '992', '255', '66', '228', '690', '676', '1868', '216', '90', '993', '1649',
#                                 '688', '256', '380', '971', '44', '1', '1', '598', '998', '678', '379', '58', '84', '1284', '1340', '681', '212', '967', '260', '263']



# @handle_exceptions
# def excel_data_import_party(request):
#     # file = request.FILES.get('file')
#     # if not file:
#     #     res = 'No file was provided'
#     #     return get(res, INVALID_ERROR)

#     # this code for mobile codes ----------------------------------------
#     # try:
#     #     file ="/home/ayushgurjar/Documents/TransportSimple/transportsimple-server/transportsimple/base/data_dump/countries.json"
#     #     with open(file,'r') as f:
#     #         codes = json.load(f)
#     #         list = []
#     #         for data in codes:
#     #             code = data['phone_code'].replace('+','').replace('-','')
#     #             list.append(code)
#     #         mobile_codes = list
#     # except Exception as e:
#     #     res = f'Error reading the mobile codes file {str(e)}'
#     #     return get(res, INVALID_ERROR)
    
#     #--------------------------------------------------------------------
#     tenant = request.tenant
#     company = tenant.get_company
#     user = TenantUser.objects.filter(is_super_admin=True, tenant=tenant).first().user
    

#     file_path = '/home/ayushgurjar/Downloads/DSA-01B/Party_Import.xlsx'
#     if not os.path.exists(file_path):
#         res = "File not found at the specified path"
#         return get(res, INVALID_ERROR)
    
#     try:
#         # workbook = load_workbook(file, read_only=True)
#         workbook = load_workbook(filename=file_path, read_only=True)
#         sheet = workbook.active
#     except Exception as e:
#         res = f'Error reading the Excel file {str(e)}'
#         return get(res, INVALID_ERROR)
    
#     parties = []
#     try:
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if row[0] == None:
#                 break

#             # this code for party type--------------------------
#             party_type = row[0]
#             if party_type.lower() == "client":
#                 party_type = 0
#             elif party_type.lower() == "vendor":
#                 party_type = 1
#             else:
#                 party_type = -1

#             #this code for vender party type-------------------
#             vender_type = row[1]
#             if vender_type.lower() == "general":
#                 vender_type = 0
#             elif vender_type.lower() == "fuel":
#                 vender_type = 1
#             else:
#                 vender_type = -1

#             ##-------this code for mobile code and number-------
#             m_code = "971"
#             if row[6]==None:
#                 mob = ''
#                 m_code = ''
#             if type(row[6])==float:
#                 mob = str(int(row[6]))
#             if type(row[6])==int:
#                 mob=str(row[6])
#             if type(row[6])==str:
#                 mob = row[6].replace('=','').replace('+','').replace('-','').replace(' ','')
#                 mob = mob.split('/')[0].split('&')[0] 
#             if mob[:2]=='00':
#                 mob=mob[2:]
#             code = mob[:3]
#             if code in mobile_codes:
#                 m_code,mob=code,mob[3:]
#             if m_code != '':     
#                 m_code = "+"+m_code

#             #---------this code for gmail --------------
#             gmail=''
#             if row[3]!=None:
#                 gmail = row[3]
#                 gmail=gmail.lower()
#                 if gmail=="@gmail.com":
#                     gmail = ''

#             #------------this code for TRN-----------
#             if row[8]==None:
#                 trn = ''
#             elif type(row[8])==str:
#                 trn = row[8]
#                 if trn.isalnum():
#                     trn = trn
#                 else:
#                     trn = trn.replace('/','').split()
#                     for i in trn:
#                         if i.isalnum():
#                             trn = i
#                         else:
#                             trn=''
#                 if trn.isalpha():
#                     trn=''
#             elif type(row[8])==int:
#                 trn = str(row[8])
#             elif type(row[8])==float:
#                 trn = str(int(row[8]))
#             if len(trn)<=10 and trn!="":
#                 trn = ""
#             if len(trn)>15:
#                 trn = ""

#             data = {
#                 'party_type': party_type,
#                 'vendor_party_type' : vender_type,
#                 'company_name': row[2],                              
#                 'email_address' : gmail,                                 
#                 'display_name' : row[4],                             
#                 'mobile_country_code' : m_code,                  
#                 'mobile' : mob,                           
#                 'VAT Registration' : row[7] if row[7] else '',  # not in use
#                 'TRN' : trn
#             }
#             parties.append(data)

#     except Exception as e:
#         res = f'Error importing data into the database: {str(e)}'
#         return get(res, INVALID_ERROR)
    
#     with transaction.atomic():
#         try:
#             for party in parties:
#                 partytype = party['party_type']
#                 vendor_party_type = party['vendor_party_type']
#                 company_name = party['company_name']
#                 email_address = party['email_address']
#                 display_name = party['display_name']
#                 mobile = party['mobile']
#                 foreman_monthly_pay = 0
#                 mobile_country_code = party['mobile_country_code']
#                 trn_no = party['TRN']

#                 # create party-------------------------------------------------
#                 party_exists = Party.objects.filter(tenant=tenant, company_name=company_name).exists()
#                 if party_exists:
#                     continue
#                 party_saved_obj = Party.objects.create(**{"tenant":request.tenant,"party_type":partytype,
#                                              "vendor_party_type":vendor_party_type,
#                                              "company_name":company_name,"email_address":email_address,
#                                              "display_name":display_name,"mobile":mobile,
#                                              "foreman_monthly_pay":foreman_monthly_pay,
#                                              "mobile_country_code":mobile_country_code,"created_by":user})


#                 party_content_type = ContentType.objects.get_for_model(Party)

#                 party_context = {
#                     "tenant": request.tenant,
#                     "object_id": party_saved_obj.id,
#                     "content_type": party_content_type,
#                     'user': user
#                     }
                
#                 # #address create----------------------------------------------------------------------------
#                 Address.objects.create(tenant=request.tenant,
#                     content_type=party_content_type,
#                     object_id=party_saved_obj.id, address_type=0)
#                 Address.objects.create(tenant=request.tenant,
#                     content_type=party_content_type,
#                     object_id=party_saved_obj.id, address_type=1)
           
#                 #---------------contact create-------------------------------------------------
#                 PartyContact.objects.create(tenant=request.tenant,party_id=party_saved_obj.id)
     
#                 # # tax create-----------------------------------------------------------------------------
#                 GST_UNREG = "aa5ce0dc-bf9f-4e01-b7bd-ed58b76d9ac4"
#                 GST_REG   = "e8f46694-0868-4ad5-8f13-ba32abb55522"
#                 tax_details = {
#                                "apply_tds": False,
#                                "gstin": trn_no,
#                                "id": None,
#                                "pan": "",
#                                "tds_attachment": [],
#                                "tds_declaration": False,
#                                "treatment": GST_UNREG if trn_no == "" else GST_REG}

#                 tax_op_status = party_tax_creator(party_saved_obj, party_context,
#                                           tax_details, company)
#                 if tax_op_status[0] == "error":
#                     transaction.set_rollback(True)
                

#         except Exception as e:
#             transaction.set_rollback(True)
#             res = f'Error importing data into the database: {str(e)}'
#             return get(res, SUCCESS)
        
#     res = f"Data imported successfully"
#     return get(res, SUCCESS)

                         ##################################################party########################################


# from django.db import transaction
# from transportsimple.base.exception_handler import FormError, InvalidDataError, handle_exceptions
# from transportsimple.base.service_status import created, deleted, updated, get
# from transportsimple.vehicles.repo.details.models import Vehicle
# import os
# from datetime import datetime
# from django.db import transaction
# from openpyxl import load_workbook
# from transportsimple.base.messages import * 
# from transportsimple.vehicles.repo.makemodel.models import VehicleBrand, VehicleMake,VehicleModel
# from transportsimple.vehicles.services.detail.utils import get_num_of_vehicles, \
#     create_tyre_for_existing_model, update_vehicle_odometer
# from transportsimple.vehicles.utils import create_default_docs
# from transportsimple.vehicles.repo.document.models import VehicleWarranty, \
#     VehicleDocument
# from transportsimple.company.repo.misc.models import DropdownOptionValues


# @handle_exceptions
# def excel_data_import_vehicle(request):
#     # file = request.FILES.get('file')
#     # if not file:
#     #     res = 'No file was provided'
#     #     return get(res, INVALID_ERROR)
    

#     file_path = '/home/ayushgurjar/Downloads/DSA-01B/Vehicle_Import.xlsx'
#     if not os.path.exists(file_path):
#         res = "File not found at the specified path"
#         return get(res, INVALID_ERROR)

#     try:
#         # workbook = load_workbook(file, read_only=True)
#         workbook = load_workbook(filename=file_path, read_only=True)
#         sheet = workbook.active
#     except Exception as e:
#         res = f'Error reading the Excel file {str(e)}'
#         return get(res, INVALID_ERROR)

#     vehicles = []
#     try:
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             if row[0] == None:
#                 break
#             #------this code for vehicle num,type----------
#             mix = row[4].split()
#             num = mix[0]
#             v_type = " ".join(mix[1:])

#             #----------------this code for Vehicle Make Year----------------------
#             make_year = str(int(row[6]))

#             #----------------this code for Registration Expiry Date----------------------
#             if type(row[7])==str:
#                 exp_date = row[7].replace("-","/")
#                 if len(exp_date) == 10:
#                     exp_date = datetime.strptime(exp_date, '%d/%m/%Y').date()
#                 else:
#                     exp_date = datetime.strptime(exp_date, '%d/%m/%y').date()
#             else:
#                 exp_date= (row[7]).date()

#             #----------------this code for Registration Date----------------------           
#             r_date = (row[8])
#             if type(r_date)==str:
#                 r_date = r_date.replace("-","/")
#                 if len(r_date) == 10:
#                     r_date = datetime.strptime(r_date, '%d/%m/%Y').date()
#                 else:
#                     r_date = datetime.strptime(r_date, '%d/%m/%y').date()
#             else:
#                 r_date= r_date.date()
#             if not r_date<exp_date:
#                 r_date,exp_date = exp_date,r_date

#             data = {
#                     'owner_type' : row[0],               
#                     'Company Name' : row[1],                              
#                     'Vehicle Make' : row[2],    
#                     'Vehicle Model' : row[3] if row[3] else '',                             
#                     'reg_number' : num,   
#                     'vehicle_type': v_type,                        
#                     'chassis_number' : row[5],                  
#                     'Vehicle Make_Year' : make_year,
#                     'issue_date' : r_date,       
#                     'expiry_date' : exp_date,          
#             }
#             vehicles.append(data)
       
#     except Exception as e:
#         res = f'Error importing data into the database: {str(e)}'
#         return get(res, INVALID_ERROR)

#     #-------data-saving-logic---------
#     with transaction.atomic():
#         try:
#             for veh in vehicles:
#                 owner = veh["Company Name"]
#                 chassis_number = veh["chassis_number"]
#                 reg_number = veh["reg_number"]
#                 make_name = veh["Vehicle Make"]
#                 model_name= veh["Vehicle Model"]
#                 label_v = veh["vehicle_type"]
#                 label_o = veh["owner_type"] #not in use
#                 issue_date = veh["issue_date"]
#                 expiry_date = veh["expiry_date"]

#                 # check vehicle is saved ----------------------------------------
#                 reg_number = reg_number
#                 check_reg_no = Vehicle.objects.check_vehicle_exists_by_reg_number(
#                             tenant=request.tenant, reg_number=reg_number)
#                 if check_reg_no:
#                     continue
                
#                 # create make-----------------------------
#                 obj_make, _i = VehicleMake.objects.get_or_create(tenant=request.tenant,make_name=make_name,
#                                                         defaults={"tenant":request.tenant,
#                                                             "make_name":make_name})

#                 #  craete model--------------------------------------
#                 obj_model, _j = VehicleModel.objects.get_or_create(tenant=request.tenant,
#                                                                 model_name=model_name,
#                                                                 vehicle_make=obj_make,
#                                                                 defaults={"tenant":request.tenant,
#                                                                 "model_name":model_name,
#                                                                 "vehicle_make":obj_make})

#                 #create brand ----------------------------------------------
#                 obj_brand,_b = VehicleBrand.objects.get_or_create(tenant=request.tenant,
#                                                             created_by=request.user,
#                                                             make = obj_make.id,
#                                                                 model = obj_model.id,
#                                                                 defaults={"tenant" : request.tenant,
#                                                             "make" : obj_make,
#                                                                 "model" : obj_model})


#                 # create vehicle_type----------------------------------
#                 obj_v_type,_v = DropdownOptionValues.objects.get_or_create(tenant=request.tenant,
#                                                             label=label_v,
#                                                             defaults={"tenant":request.tenant,
#                                                             "key":"vehicle-type","label":label_v})

#                 # create vehicle ------------------------------------------------
#                 obj_vehicle = Vehicle.objects.create(tenant=request.tenant,
#                                                     owner=owner,chassis_number=chassis_number,reg_number=reg_number,
#                                                     brand=obj_brand,vehicle_type=obj_v_type) #,owner_type=obj_o_type
                
#                 create_tyre_for_existing_model(obj_vehicle)
#                 create_default_docs(obj_vehicle)

#                 # update issue date and expiry date ----------------------------------------
#                 obj_doc = VehicleDocument.objects.get(tenant=request.tenant,vehicle=obj_vehicle.id,name = "Registration Certificate")
#                 obj_doc.issue_date = issue_date
#                 obj_doc.expiry_date = expiry_date
#                 obj_doc.save()
  
#         except Exception as e:
#             transaction.set_rollback(True)
#             res = f'Error importing data into the database: {str(e)}'
#             return get(res, INVALID_ERROR)

#     res = f"Data imported successfully "
#     return get(res, SUCCESS)


##############################################################  end   ##############################################################


################################## bdp milestone ################################

ms_data = {
          "Event" : {
            "feature" : "BOTH", # Always BOTH
            "dray_Scac" : "PRIM", # Always PRIM
            "empty_Indicator" : "Ful", # E - Empty/F - Full
            "carrier_Scac_Code" : "HLCU", # carrierScacCode
            "container_Number" : "TBNU01", # containerNumber PRIME LINK will enter
            "booking_Number" : "3645546646", # bookingNumber
            "shipment_Number" : "0042239933", # shipmentReferenceNumber
            "container_Pool_Id" : "", # containerNumber -- doubt
            "container_Type_Code" : "4010", # containerType
            "tender_Request_Number" : 23071800001, # tenderRequestNumber
            "shipper_Name" : "", # Not Required
            "milestone" : "Full Delivery", # AF - Empty Pickup / X8 - Empty Drop /CP - Full Pickup /I1 - Full Delivery
            "milestone_Timestamp" : "2023-Aug-05 10:25:00",
            "milestone_Time" : "10:25",
            "milestone_Location" : "DP World" 
          }
        }

data = ms_data["Event"]

Indicator=data["empty_Indicator"]
if Indicator == "Empty":
    Indicator = "E"
elif Indicator == "Full":
    Indicator = "F"
else:
    Indicator = ""
data["empty_Indicator"]=Indicator


# milestone= "empty Drop"
milestone = data["milestone"]
if milestone == "Empty Pickup":
    milestone = "AF"
elif milestone == "Empty Drop":
    milestone = "X8"
elif milestone == "Full Pickup":
    milestone = "CP"
elif milestone == "Full Delivery":
    milestone = "I1"
else:
    milestone = ""
data["milestone"]=milestone


# milestone= "empty Drop"
# try:
#     match milestone:
#         case"Empty Pickup":
#             milestone = "AF"
#         case "Empty Drop":
#             milestone = "X8"
#         case "Full Pickup":
#             milestone = "CP"
#         case "Full Delivery":
#             milestone = "I1"
# except:
#     print("err")
# data["milestone"]=milestone


ms_data = {
          "Event" : {
            "feature" : "BOTH",
            "drayScac" : "PRIM",
            "emptyIndicator" : data["empty_Indicator"],
            "carrierScacCode" : data["carrier_Scac_Code"],
            "containerNumber" : data["container_Number"],
            "bookingNumber" : data["booking_Number"],
            "shipmentNumber" : data["shipment_Number"],
            "containerPoolId" : data["container_Pool_Id"],
            "containerTypeCode" : data["container_Type_Code"],
            "tenderRequestNumber" : data["tender_Request_Number"],
            "shipperName" : data["shipper_Name"],
            "milestone" : data["milestone"],
            "milestoneTimestamp" : data["milestone_Timestamp"],
            "milestoneTime" : data["milestone_Time"],
            "milestoneLocation" : data["milestone_Location"]
          }
        }
# print(ms_data)



# {
#         "feature" : "BOTH",
#         "dray_Scac" : "PRIM",
#         "empty_Indicator" : "Empty",
#         "carrier_Scac_Code" : "HLCU",
#         "container_Number" : "TBNU01",
#         "booking_Number" : "3645546646",
#         "shipment_Number" : "0042239933",
#         "container_Pool_Id" : "",
#         "container_Type_Code" : "4010",
#         "tender_Request_Number" : 23071800001,
#         "shipper_Name" : "",
#         "milestone" : "Full Delivery",
#         "milestone_Timestamp" : "2023-Aug-05 10:25:00",
#         "milestone_Time" : "10:25",
#         "milestone_Location" : "DP World" 
# }


    # data = {
    #     "feature" : data['feature'],
    #     "dray_Scac" : data['dray_Scac'],
    #     "empty_Indicator" : data['empty_Indicator'],
    #     "carrier_Scac_Code" : data['carrier_Scac_Code'],
    #     "container_Number" : data['container_Number'],
    #     "booking_Number" : data['booking_Number'],
    #     "shipment_Number" : data['shipment_Number'],
    #     "container_Pool_Id" : data['container_Pool_Id'],
    #     "container_Type_Code" : data['container_Type_Code'],
    #     "tender_Request_Number" : data['tender_Request_Number'],
    #     "shipper_Name" : data['shipper_Name'],
    #     "milestone" : data['milestone'],
    #     "milestone_Timestamp" : data['milestone_Timestamp'],
    #     "milestone_Time" : data['milestone_Time'],
    #     "milestone_Location" : data['milestone_Location'] }






def excel_data_import3():
    file_path = '/home/ayushgurjar/Downloads/DSA-01B/plant to cp code mapping.xlsx'

    if not os.path.exists(file_path):
        print("File not found at the specified path")

    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active
        # print(sheet)
    except Exception as e:
        print(f'Error reading the Excel file {str(e)}')

    file = {}
    total = 0
    try:
        for row in sheet.iter_rows(min_row=2,values_only=True):
            if row[1] == None:
                continue
            # print(row[1],"----",row[4])
           
            file[row[1]] = row[4]
            total+=1
    except Exception as e:
        print(f'Error importing data into the database: {str(e)}')

    print(file)
    # filename = "cp code mapping.json"
    # with open(filename, "w") as f:
    #     json.dump(file, f, indent=4)

# excel_data_import3()
# uvicorn main:app --host 127.0.0.1 --port 3000

    # data = {
    #     "feature" : data['feature'],    -
    #     "dray_Scac" : data['dray_Scac'],   -
    #     "empty_Indicator" : data['empty_Indicator'],   - 
    #     "carrier_Scac_Code" : data['carrier_Scac_Code'],  - 
    #     "container_Number" : data['container_Number'],   
    #     "booking_Number" : data['booking_Number'],     
    #     "shipment_Number" : data['shipment_Number'],
    #     "container_Pool_Id" : data['container_Pool_Id'],
    #     "container_Type_Code" : data['container_Type_Code'],
    #     "tender_Request_Number" : data['tender_Request_Number'],
    #     "shipper_Name" : data['shipper_Name'],
    #     "milestone" : data['milestone'],
    #     "milestone_Timestamp" : data['milestone_Timestamp'],
    #     "milestone_Time" : data['milestone_Time'],
    #     "milestone_Location" : data['milestone_Location'] }

# {
#         "feature" : "BOTH",     -
#         "dray_Scac" : "PRIM",    -
#         "empty_Indicator" : "Empty",      -?
#         "carrier_Scac_Code" : "HLCU",      -
#         "container_Number" : "TBNU01",      - or -u
#         "booking_Number" : "3645546646",    -
#         "shipment_Number" : "0042239933",   -  
#         "container_Pool_Id" : "",           - 
#         "container_Type_Code" : "4010",     -
#         "tender_Request_Number" : 23071800001,  --
#         "shipper_Name" : "",  -u
#         "milestone" : "Full Delivery",   -u
#         "milestone_Timestamp" : "2023-Aug-05 10:25:00",  -u
#         "milestone_Time" : "10:25",   ?
#         "milestone_Location" : "DP World"  ??
# }

# "code": "AE_EXDOWDXB_GF63",
# "name": "HORIZON TERMINALS",  ?

# "shipmentReferenceNumber": "0042241187"
# "carrierScacCode": "ONEY"
# "tenderRequestNumber": 23062200040,
# "bookingNumber": "DXBD00863500",
# "containerType": 2000,

# "containerNumber": "TBNU01",
# data = {}
# tender =1 #await db['tender_details'].find_one({"tenderRequestNumber":"tender number"})

# codes = {'C092': 'CP_000530', 'C093': 'CP_000530', 'C094': 'CP_000530', 'CC12': 'CP_000538',
#           'CC13': 'CP_000538', 'CC14': 'CP_000538', 'GH10': 'CP_000540', 'GH11': 'CP_000540',
#             'GH12': 'CP_000540', 'C095': 'CP_000554', 'C096': 'CP_000554', 'C097': 'CP_000554',
#               'GF61': 'CP_000556', 'GF62': 'CP_000556', 'GF63': 'CP_000556', 'C541': 'CP_000558',
#                 'C542': 'CP_000558', 'C543': 'CP_000558', 'CD14': 'CP_000560', 'CD15': 'CP_000560',
#                   'CD16': 'CP_000560'}
# codes = {'C092': 'CP_000530', 'C093': 'CP_000530', 'C094': 'CP_000530', 'CC12': 'CP_000538',
#           'CC13': 'CP_000538', 'CC14': 'CP_000538', 'GH10': 'CP_000540', 'GH11': 'CP_000540',
#             'GH12': 'CP_000540', 'C095': 'CP_000554', 'C096': 'CP_000554', 'C097': 'CP_000554',
#               'GF61': 'CP_000556', 'GF62': 'CP_000556', 'GF63': 'CP_000556', 'C541': 'CP_000558',
#                 'C542': 'CP_000558', 'C543': 'CP_000558', 'CD14': 'CP_000560', 'CD15': 'CP_000560',
#                   'CD16': 'CP_000560'}
# code = tender['originAddress']['code'].split("_")

# data['container_Pool_Id'] = codes[code[2]]
# data['milestone_Location'] = tender['originAddress']['name']  #  ????
# data["carrier_Scac_Code"] = tender['carrierScacCode']
# data['shipment_Number'] = tender['shipmentReferenceNumber'] 
# data["container_Number"] = tender['containerNumber'] #--u
# data["booking_Number"] = tender['bookingNumber']
# data['container_Type_Code'] = tender['containerType']
# # data['tender_Request_Number'] = tender_number
# new = data['milestone_Time'].split(" ")[1].split(":")
# data['milestone_Time'] =  new[0]+":"+new[1]


res = {}
res['invoice'] = {'due_percent' : 2562}
res['invoice'].update({'over_due_percent' : 100})

total = res['invoice']['due_percent'] + res['invoice']['over_due_percent']
if total > 0:
    res['invoice']['due_percent'] = round((res['invoice']['due_percent']/total)*100, 2)
    print(100-res['invoice']['due_percent'])
    res['invoice']['over_due_percent'] = round(100 - res['invoice']['due_percent'], 2)

print(res,total)

# v1/party_portal/dashboard/